#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
실제 배분 결과 데이터를 엑셀로 변환하는 스크립트
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def try_read_csv_with_encodings(file_path):
    """여러 인코딩을 시도하여 CSV 파일을 읽습니다"""
    encodings = ['utf-8', 'cp949', 'euc-kr', 'latin1', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding)
            print(f"성공적으로 읽음: {file_path} (인코딩: {encoding})")
            return df
        except UnicodeDecodeError:
            print(f"실패: {file_path} (인코딩: {encoding})")
            continue
        except Exception as e:
            print(f"기타 오류: {file_path} (인코딩: {encoding}) - {e}")
            continue
    
    raise ValueError(f"모든 인코딩 시도 실패: {file_path}")

def load_real_data():
    """실제 배분 데이터와 매장 정보를 로드"""
    # 실제 배분 데이터 로드
    # df_real = try_read_csv_with_encodings('data/data_real_25s.csv')
    df_real = try_read_csv_with_encodings('data/data_real_24f.csv')
    
    # 매장 정보 로드
    # df_shop = try_read_csv_with_encodings('data/shop_real_control_25s.csv')
    df_shop = try_read_csv_with_encodings('data/shop_real_control_24f.csv')
    
    # 매장 이름 딕셔너리 생성
    shop_names = dict(zip(df_shop['SHOP_ID'].astype(str), df_shop['SHOP_NM_SHORT']))
    
    # QTY_SUM 정보 딕셔너리 생성
    qty_sum_dict = dict(zip(df_shop['SHOP_ID'].astype(str), df_shop['QTY_SUM']))
    
    return df_real, shop_names, qty_sum_dict

def create_sku_column(row):
    """SKU 컬럼 생성 (COLOR_CD + SIZ_CD)"""
    return row['COLOR_CD'] + str(row['SIZ_CD'])

def process_real_data(df_real):
    """실제 배분 데이터 처리"""
    # SKU 컬럼 생성
    df_real['SKU'] = df_real.apply(create_sku_column, axis=1)
    df_real['TO_LGT_SHOP_CD'] = df_real['TO_LGT_SHOP_CD'].astype(str)
    
    # 스타일별로 그룹화하여 처리
    grouped_data = {}
    for style in df_real['PROD_CD'].unique():
        style_data = df_real[df_real['PROD_CD'] == style]
        
        # 매장별, SKU별 수량 집계
        allocation_matrix = style_data.groupby(['TO_LGT_SHOP_CD', 'SKU'])['QTY'].sum().reset_index()
        allocation_pivot = allocation_matrix.pivot(index='TO_LGT_SHOP_CD', columns='SKU', values='QTY').fillna(0)
        
        grouped_data[style] = {
            'allocation_matrix': allocation_pivot,
            'raw_data': style_data
        }
    
    return grouped_data

def get_size_sort_key(size):
    """사이즈 정렬을 위한 키 생성 - 일반적인 사이즈 순서 적용"""
    # 텍스트 사이즈 순서
    text_size_order = {'XS': 0, 'S': 1, 'M': 2, 'L': 3, 'XL': 4, 'XXL': 5}
    
    if size in text_size_order:
        return (0, text_size_order[size])
    
    # 숫자 사이즈 처리 (85, 90, 95, 100, 105 등)
    try:
        numeric_size = int(size)
        return (1, numeric_size)  # 숫자는 텍스트보다 뒤에 오도록
    except:
        return (2, size)  # 기타는 맨 뒤

def create_sku_mapping(raw_data):
    """SKU별 색상/사이즈 매핑 생성"""
    sku_mapping = {}
    for _, row in raw_data.iterrows():
        sku = row['COLOR_CD'] + str(row['SIZ_CD'])
        sku_mapping[sku] = {
            'color': row['COLOR_CD'],
            'size': str(row['SIZ_CD'])
        }
    return sku_mapping

def create_real_data_excel(style_code, allocation_matrix, shop_names, qty_sum_dict, save_path, raw_data=None):
    """실제 배분 데이터를 엑셀로 변환"""
    
    # SKU별 색상/사이즈 매핑 생성
    sku_mapping = create_sku_mapping(raw_data) if raw_data is not None else {}
    
    # 워크북 생성
    wb = Workbook()
    
    # 기본 시트 삭제
    wb.remove(wb.active)
    
    # 매장 정보 준비 (QTY_SUM 기준 내림차순 정렬)
    stores_with_qty = []
    for store_id in allocation_matrix.index:
        if store_id in shop_names and store_id in qty_sum_dict:
            shop_name = shop_names[store_id]
            qty_sum = qty_sum_dict[store_id]
            stores_with_qty.append((store_id, shop_name, qty_sum))
    
    stores_with_qty.sort(key=lambda x: x[2], reverse=True)
    
    # SKU 목록
    skus = allocation_matrix.columns.tolist()
    
    # SKU를 색상과 사이즈로 분리하여 정렬 (매핑 사용)
    sku_info = []
    for sku in skus:
        if sku in sku_mapping:
            # 매핑된 정보 사용
            color = sku_mapping[sku]['color']
            size = sku_mapping[sku]['size']
        else:
            # 기존 로직 (fallback)
            if len(sku) >= 2:
                if sku.endswith(('XS', 'XL', 'XXL')):
                    color = sku[:-2]
                    size = sku[-2:]
                else:
                    color = sku[:-1]
                    size = sku[-1:]
            else:
                color = sku
                size = ''
        
        sku_info.append((sku, color, size))
    
    # 색상별, 사이즈별 정렬 (수정된 사이즈 정렬 로직 적용)
    sku_info.sort(key=lambda x: (x[1], get_size_sort_key(x[2])))
    sorted_skus = [item[0] for item in sku_info]
    
    # 1. 배분 매트릭스 시트
    ws_matrix = wb.create_sheet("배분_매트릭스")
    
    # 헤더 작성
    ws_matrix.cell(1, 1, "매장명")
    for col_idx, sku in enumerate(sorted_skus, 2):
        ws_matrix.cell(1, col_idx, sku)
    
    # 데이터 작성
    for row_idx, (store_id, shop_name, qty_sum) in enumerate(stores_with_qty, 2):
        # 매장명 (QTY 정보 포함)
        store_label = f"{shop_name} (QTY:{qty_sum:,})"
        ws_matrix.cell(row_idx, 1, store_label)
        
        # 각 SKU별 배분 수량
        for col_idx, sku in enumerate(sorted_skus, 2):
            if store_id in allocation_matrix.index and sku in allocation_matrix.columns:
                qty = allocation_matrix.loc[store_id, sku]
                if qty > 0:
                    ws_matrix.cell(row_idx, col_idx, int(qty))
    
    # 헤더 스타일 적용
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col in range(1, len(sorted_skus) + 2):
        cell = ws_matrix.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 2. 매장별 통계 시트
    ws_store_stats = wb.create_sheet("매장별_통계")
    
    store_stats = []
    for store_id, shop_name, qty_sum in stores_with_qty:
        if store_id in allocation_matrix.index:
            store_data = allocation_matrix.loc[store_id]
            allocated_skus = (store_data > 0).sum()
            total_allocated = store_data.sum()
            
            # 색상 커버리지 계산 (매핑 사용)
            colors = set()
            sizes = set()
            for sku in store_data[store_data > 0].index:
                if sku in sku_mapping:
                    # 매핑된 정보 사용
                    colors.add(sku_mapping[sku]['color'])
                    sizes.add(sku_mapping[sku]['size'])
                else:
                    # 기존 로직 (fallback)
                    if len(sku) >= 2:
                        if sku.endswith(('XS', 'XL', 'XXL')):
                            colors.add(sku[:-2])
                            sizes.add(sku[-2:])
                        else:
                            colors.add(sku[:-1])
                            sizes.add(sku[-1:])
            
            color_coverage = len(colors)
            size_coverage = len(sizes)
            
            # 전체 색상/사이즈 개수 계산 (매핑 사용)
            all_colors = set()
            all_sizes = set()
            for sku in sorted_skus:
                if sku in sku_mapping:
                    all_colors.add(sku_mapping[sku]['color'])
                    all_sizes.add(sku_mapping[sku]['size'])
                else:
                    # 기존 로직 (fallback)
                    if len(sku) >= 2:
                        if sku.endswith(('XS', 'XL', 'XXL')):
                            all_colors.add(sku[:-2])
                            all_sizes.add(sku[-2:])
                        else:
                            all_colors.add(sku[:-1])
                            all_sizes.add(sku[-1:])
            
            total_colors = len(all_colors)
            total_sizes = len(all_sizes)
            
            # 퍼센티지 계산
            color_coverage_pct = (color_coverage / total_colors * 100) if total_colors > 0 else 0
            size_coverage_pct = (size_coverage / total_sizes * 100) if total_sizes > 0 else 0
            
            store_stats.append({
                '매장명': f"{shop_name} (QTY:{qty_sum:,})",
                '총_배분수량': int(total_allocated),
                '배분된_SKU_수': allocated_skus,
                '색상_커버리지': f"{color_coverage_pct:.1f}%",
                '사이즈_커버리지': f"{size_coverage_pct:.1f}%"
            })
    
    df_store_stats = pd.DataFrame(store_stats)
    
    # 데이터프레임을 워크시트에 추가
    for r in dataframe_to_rows(df_store_stats, index=False, header=True):
        ws_store_stats.append(r)
    
    # 헤더 스타일 적용
    for col in range(1, len(df_store_stats.columns) + 1):
        cell = ws_store_stats.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 3. SKU별 통계 시트
    ws_sku_stats = wb.create_sheet("SKU별_통계")
    
    sku_stats = []
    for sku in sorted_skus:
        if sku in allocation_matrix.columns:
            sku_data = allocation_matrix[sku]
            allocated_stores = (sku_data > 0).sum()
            total_allocated = sku_data.sum()
            
            # 색상과 사이즈 분리 (매핑 사용)
            if sku in sku_mapping:
                color = sku_mapping[sku]['color']
                size = sku_mapping[sku]['size']
            else:
                # 기존 로직 (fallback)
                if len(sku) >= 2:
                    if sku.endswith(('XS', 'XL', 'XXL')):
                        color = sku[:-2]
                        size = sku[-2:]
                    else:
                        color = sku[:-1]
                        size = sku[-1:]
                else:
                    color = sku
                    size = ''
            
            sku_stats.append({
                'SKU': sku,
                '색상': color,
                '사이즈': size,
                '배분된_매장_수': allocated_stores,
                '총_배분량': int(total_allocated)
            })
    
    df_sku_stats = pd.DataFrame(sku_stats)
    
    # 데이터프레임을 워크시트에 추가
    for r in dataframe_to_rows(df_sku_stats, index=False, header=True):
        ws_sku_stats.append(r)
    
    # 헤더 스타일 적용
    for col in range(1, len(df_sku_stats.columns) + 1):
        cell = ws_sku_stats.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 4. 요약 시트
    ws_summary = wb.create_sheet("요약")
    
    # 요약 정보 계산
    total_stores = len(stores_with_qty)
    total_skus = len(sorted_skus)
    total_allocated_qty = allocation_matrix.sum().sum()
    
    # 평균 커버리지 계산
    avg_color_coverage = df_store_stats['색상_커버리지'].str.rstrip('%').astype(float).mean()
    avg_size_coverage = df_store_stats['사이즈_커버리지'].str.rstrip('%').astype(float).mean()
    avg_filled_cells = df_store_stats['배분된_SKU_수'].mean()
    
    summary_data = [
        ['항목', '값'],
        ['스타일_코드', style_code],
        ['총_매장_수', total_stores],
        ['총_SKU_수', total_skus],
        ['총_배분_수량', int(total_allocated_qty)],
        ['평균_색상_커버리지', f"{avg_color_coverage:.2f}%"],
        ['평균_사이즈_커버리지', f"{avg_size_coverage:.2f}%"],
        ['평균_채워진_셀_개수', f"{avg_filled_cells:.2f}"],
        ['생성_시간', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # 헤더 스타일 적용
    for col in range(1, 3):
        cell = ws_summary.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 파일 저장
    wb.save(save_path)
    print(f"실제 배분 데이터 엑셀 파일이 생성되었습니다: {save_path}")
    
    return {
        'total_stores': total_stores,
        'total_skus': total_skus,
        'total_allocated_qty': int(total_allocated_qty)
    }

def main():
    """메인 함수"""
    print("실제 배분 데이터 엑셀 생성 시작...")
    
    # 데이터 로드
    df_real, shop_names, qty_sum_dict = load_real_data()
    
    # 데이터 처리
    grouped_data = process_real_data(df_real)
    
    # 출력 디렉토리 생성
    os.makedirs('output', exist_ok=True)
    
    # 각 스타일별로 엑셀 파일 생성
    for style_code, data in grouped_data.items():
        print(f"스타일 {style_code} 처리 중...")
        
        # 파일명 생성 (타임스탬프 추가)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"실제배분_{style_code}_allocation_matrix_{timestamp}.xlsx"
        save_path = os.path.join('output', filename)
        
        # 엑셀 파일 생성
        result = create_real_data_excel(
            style_code, 
            data['allocation_matrix'], 
            shop_names, 
            qty_sum_dict, 
            save_path,
            data['raw_data']  # raw_data 전달
        )
        
        print(f"  - 매장: {result['total_stores']}개")
        print(f"  - SKU: {result['total_skus']}개")
        print(f"  - 총 배분량: {result['total_allocated_qty']}개")
    
    print("모든 실제 배분 데이터 엑셀 파일 생성 완료!")

if __name__ == "__main__":
    main() 